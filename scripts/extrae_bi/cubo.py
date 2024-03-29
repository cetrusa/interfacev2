import os
import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
import logging
from scripts.StaticPage import StaticPage
from scripts.conexion import Conexion as con
from scripts.config import ConfigBasic
import ast
import xlsxwriter

# Configuración del logging
logging.basicConfig(
    filename="logCubo.txt",
    level=logging.DEBUG,
    format="%(asctime)s %(message)s",
    filemode="w",
)


class DataBaseConnection:
    def __init__(self, config, mysql_engine=None, sqlite_engine=None):
        self.config = config
        # Asegurarse de que los engines son instancias de conexión válidas y no cadenas
        self.engine_mysql = mysql_engine if mysql_engine else self.create_engine_mysql()
        print(self.engine_mysql)
        self.engine_sqlite = (
            sqlite_engine if sqlite_engine else create_engine("sqlite:///mydata.db")
        )
        print(self.engine_sqlite)

    def create_engine_mysql(self):
        # Simplificación en la obtención de los parámetros de configuración
        user, password, host, port, database = (
            self.config.get("nmUsrIn"),
            self.config.get("txPassIn"),
            self.config.get("hostServerIn"),
            self.config.get("portServerIn"),
            self.config.get("dbBi"),
        )
        return con.ConexionMariadb3(
            str(user), str(password), str(host), int(port), str(database)
        )

    def execute_query_mysql(self, query, chunksize=None):
        # Usar el engine correctamente
        with self.create_engine_mysql.connect() as connection:
            cursor = connection.execution_options(isolation_level="READ COMMITTED")
            return pd.read_sql_query(query, cursor, chunksize=chunksize)

    def execute_sql_sqlite(self, sql, params=None):
        # Usar el engine correctamente
        with self.engine_sqlite.connect() as connection:
            return connection.execute(sql, params)

    def execute_query_mysql_chunked(self, query, table_name, chunksize=50000):
        try:
            # Primero, elimina la tabla si ya existe
            self.eliminar_tabla_sqlite(table_name)
            # Luego, realiza la consulta y almacena los resultados en SQLite
            with self.engine_mysql.connect() as connection:
                cursor = connection.execution_options(isolation_level="READ COMMITTED")

                for chunk in pd.read_sql_query(query, con=cursor, chunksize=chunksize):
                    chunk.to_sql(
                        name=table_name,
                        con=self.engine_sqlite,
                        if_exists="append",
                        index=False,
                    )

                # Retorna el total de registros almacenados en la tabla SQLite
                with self.engine_sqlite.connect() as connection:
                    # Modificar aquí para usar fetchone correctamente
                    total_records = connection.execute(
                        text(f"SELECT COUNT(*) FROM {table_name}")
                    ).fetchone()[0]
                return total_records

        except Exception as e:
            logging.error(f"Error al ejecutar el query: {e}")
            print(f"Error al ejecutar el query: {e}")
            raise
        print("terminamos de ejecutar el query")

    def eliminar_tabla_sqlite(self, table_name):
        sql = text(f"DROP TABLE IF EXISTS {table_name}")
        with self.engine_sqlite.connect() as connection:
            connection.execute(sql)


class CuboVentas:
    def __init__(self, database_name, IdtReporteIni, IdtReporteFin):
        self.database_name = database_name
        self.IdtReporteIni = IdtReporteIni
        self.IdtReporteFin = IdtReporteFin
        self.configurar(database_name)
        self.file_path = None
        self.archivo_cubo_ventas = None

    def configurar(self, database_name):
        try:
            config_basic = ConfigBasic(database_name)
            self.config = config_basic.config
            # config_basic.print_configuration()
            self.db_connection = DataBaseConnection(config=self.config)
            self.engine_sqlite = self.db_connection.engine_sqlite
            self.engine_mysql = self.db_connection.engine_mysql
        except Exception as e:
            logging.error(f"Error al inicializar CuboVentas: {e}")
            raise

    def generate_sqlout(self, hoja):
        sql = self.config["nmProcedureExcel"]
        if self.config["dbBi"] == "powerbi_tym_eje":
            return text(
                f"CALL {sql}('{self.IdtReporteIni}','{self.IdtReporteFin}','','{str(hoja)}',0,0,0);"
            )
        return text(
            f"CALL {sql}('{self.IdtReporteIni}','{self.IdtReporteFin}','','{str(hoja)}');"
        )

    def guardar_datos(self, table_name, file_path, hoja, total_records, wb):
        if total_records > 1000000:
            self.guardar_datos_csv(table_name, file_path)
        elif total_records > 250000:
            self.guardar_datos_excel_xlsxwriter2(table_name, hoja, wb)
        else:
            self.guardar_datos_excel_xlsxwriter2(table_name, hoja, wb)

    def generar_nombre_archivo(self, hoja, ext=".xlsx"):
        self.archivo_cubo_ventas = f"Cubo_de_Ventas_{self.database_name}_de_{self.IdtReporteIni}_a_{self.IdtReporteFin}{ext}"
        self.file_path = os.path.join("media", self.archivo_cubo_ventas)
        return self.archivo_cubo_ventas, self.file_path

    def guardar_datos_csv(self, table_name, file_path):
        """
        Guarda los datos de una tabla SQLite en un archivo CSV, manejando grandes volúmenes de datos.

        Args:
        table_name (str): Nombre de la tabla en SQLite de donde se extraerán los datos.
        file_path (str): Ruta del archivo CSV de destino.

        Descripción:
        Este método extrae datos de una tabla SQLite en bloques (chunks) de tamaño definido
        y los guarda en un archivo CSV. Es útil para manejar grandes volúmenes de datos que
        no caben en la memoria de una sola vez.

        El archivo CSV se escribe en modo 'append', lo que permite añadir datos al archivo
        existente sin sobrescribirlo. Cada bloque de datos se lee de la tabla SQLite y se
        escribe en el archivo CSV hasta que todos los datos han sido procesados.

        Se utiliza un tamaño de bloque (chunksize) de 100000, lo que significa que los datos se
        leerán y escribirán en bloques de 100000 filas a la vez. Este valor puede ajustarse según
        las necesidades de rendimiento y la capacidad de la memoria.
        """
        chunksize = 50000  # Define el tamaño de cada bloque de datos

        with self.engine_sqlite.connect() as connection:
            # Iterar sobre los bloques (chunks) de datos de la tabla SQLite
            for chunk in pd.read_sql_query(
                f"SELECT * FROM {table_name}", self.engine_sqlite, chunksize=chunksize
            ):
                # Escribir el bloque de datos en el archivo CSV en modo 'append'
                chunk.to_csv(file_path, mode="a", index=False)

    def guardar_datos_excel_xlsxwriter(self, table_name, file_path, hoja):
        """
        Guarda datos de una tabla SQLite en un archivo Excel, manejando archivos grandes.

        Args:
        table_name (str): Nombre de la tabla en SQLite.
        file_path (str): Ruta del archivo Excel de destino.
        hoja (str): Nombre de la hoja de Excel.
        """
        # Definir el tamaño de cada bloque de datos (chunk)
        chunksize = 50000

        # Crear un objeto ExcelWriter con xlsxwriter como motor
        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            # Establecer conexión con la base de datos SQLite
            with self.engine_sqlite.connect() as connection:
                # Iniciar la fila desde donde empezar a escribir en el archivo Excel
                startrow = 0

                # Iterar sobre los bloques (chunks) de datos de la tabla SQLite
                for chunk in pd.read_sql_query(
                    f"SELECT * FROM {table_name}",
                    self.engine_sqlite,
                    chunksize=chunksize,
                ):
                    print(f"Procesando chunk de tamaño: {len(chunk)}")
                    # Escribir el bloque de datos en el archivo Excel
                    chunk.to_excel(
                        writer,
                        sheet_name=hoja,
                        startrow=startrow,
                        index=False,
                        header=not bool(startrow),
                    )
                    # Actualizar la fila de inicio para el siguiente bloque
                    startrow += len(chunk)

                    print(f"Próximo startrow será: {startrow}")

    def guardar_datos_excel_xlsxwriter2(self, table_name, hoja, wb):
        """
        Guarda datos de una tabla SQLite en un archivo Excel, manejando archivos grandes.

        Args:
        table_name (str): Nombre de la tabla en SQLite.
        file_path (str): Ruta del archivo Excel de destino.
        hoja (str): Nombre de la hoja de Excel.
        """
        chunksize = 50000  # Tamaño de cada bloque de datos

        # Crear un objeto ExcelWriter con xlsxwriter como motor

        ws = wb.create_sheet(title=hoja)

        # Inicialmente, extraer solo los encabezados de la tabla
        headers = pd.read_sql_query(
            f"SELECT * FROM {table_name} LIMIT 0", self.engine_sqlite
        ).columns.tolist()
        ws.append(headers)  # Agregar los encabezados a la hoja de Excel

        # Procesar los datos en chunks
        for chunk in pd.read_sql_query(
            f"SELECT * FROM {table_name}", self.engine_sqlite, chunksize=chunksize
        ):
            for index, row in chunk.iterrows():
                cells = [WriteOnlyCell(ws, value=value) for value in row]
                ws.append(cells)

    def guardar_datos_excel_openpyxl(self, table_name, file_path, hoja):
        """
        Guarda datos de una tabla SQLite en un archivo Excel, manejando archivos grandes.

        Args:
            table_name (str): Nombre de la tabla en SQLite.
            file_path (str): Ruta del archivo Excel de destino.
            hoja (str): Nombre de la hoja de Excel.
        """
        chunksize = 50000  # Tamaño de cada bloque de datos

        # Crear un libro de trabajo en modo de solo escritura
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=hoja)

        # Establecer conexión con SQLite y leer los datos en chunks
        with self.engine_sqlite.connect() as connection:
            first_chunk = True
            for chunk in pd.read_sql_table(
                table_name, con=connection, chunksize=chunksize
            ):
                # Agregar encabezados solo para el primer chunk
                if first_chunk:
                    ws.append(chunk.columns.tolist())
                    first_chunk = False

                # Escribir los datos del chunk en la hoja de Excel
                for row in chunk.itertuples(index=False, name=None):
                    ws.append(row)

        # Guardar el libro de trabajo
        wb.save(file_path)
        print(f"Archivo Excel {file_path} generado exitosamente")

        print(f"Archivo Excel {file_path} generado exitosamente")

    def guardar_datos_excel_completo(self, table_name, file_path, hoja):
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            with self.engine_sqlite.connect() as connection:
                df = pd.read_sql_table(table_name, con=connection)
                df.to_excel(writer, index=False, sheet_name=hoja)

    def procesar_hoja(self, hoja, wb):
        try:
            sqlout = self.generate_sqlout(hoja)
            print(sqlout)
            table_name = f"my_table_{self.database_name}_{hoja}"
            total_records = self.db_connection.execute_query_mysql_chunked(
                query=sqlout, table_name=table_name
            )

            print(f"Total de registros: {total_records}")

            archivo_cubo_ventas, file_path = self.generar_nombre_archivo(
                hoja, ext=".csv" if total_records > 1000000 else ".xlsx"
            )
            print(f"aqui se genero el nombre del archivo {file_path}")
            self.guardar_datos(table_name, file_path, hoja, total_records, wb)
            print(f"Archivo {archivo_cubo_ventas} generado exitosamente")

            self.db_connection.eliminar_tabla_sqlite(table_name)
            print(f"Procesamiento de la hoja {hoja} finalizado")
            return True
        except Exception as e:
            print(f"Error al procesar la hoja {hoja}: {e}")
            logging.error(f"Error al procesar la hoja {hoja}: {e}")
            return {
                "success": False,
                "error_message": f"Error al procesar la hoja {hoja}: {e}",
            }

    def procesar_datos(self):
        txProcedureExcel_str = self.config["txProcedureExcel"]
        if isinstance(txProcedureExcel_str, str):
            try:
                self.config["txProcedureExcel"] = ast.literal_eval(txProcedureExcel_str)
            except ValueError as e:
                logging.error(f"Error al convertir txProcedureExcel a lista: {e}")
                # Manejar el error o asignar un valor predeterminado
                self.config["txProcedureExcel"] = []

        if not self.config["txProcedureExcel"]:
            return {"success": False, "error_message": "No hay datos para procesar"}

        wb = Workbook(write_only=True)

        for hoja in self.config["txProcedureExcel"]:
            print(f"Procesando hoja {hoja}")
            if not self.procesar_hoja(hoja, wb):
                return {
                    "success": False,
                    "error_message": f"Error al procesar la hoja {hoja}",
                }
        # Guardar el libro de trabajo
        wb.save(self.file_path)
        print("Proceso finalizado")
        print(self.file_path)
        print(self.archivo_cubo_ventas)
        return {
            "success": True,
            "file_path": self.file_path,
            "file_name": self.archivo_cubo_ventas,
        }
